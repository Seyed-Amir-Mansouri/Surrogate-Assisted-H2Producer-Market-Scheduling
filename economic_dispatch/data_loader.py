"""Parse a zone workbook into structured pandas objects.

Each zone workbook has six sheets with a fixed schema (see README). Four are
key/value ("Parameter", "Value"); "Hourly Profiles" and "Technology
Characteristics" are tables. This module only *reads* — all modelling logic
(classification, costs) lives in model.py, except the pure name-based
``classify`` helper below which both share.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd

S_CAP = "Technology Capacities"
S_STO = "Storage Capacities"
S_RES = "Reserve Requirements"
S_PROF = "Hourly Profiles"
S_CHAR = "Technology Characteristics"
S_GH = "Gas & Hydrogen Assets"

CAT_COMMIT = "committable"
CAT_VRES = "vres"
CAT_ROR = "ror"
CAT_PROFILE = "profile_gen"
CAT_IGNORE = "ignore"


@dataclass
class ZoneData:
    code: str
    capacities: dict[str, float]
    storage_energy: dict[str, float]
    reserves: dict[str, float]
    h2_assets: dict[str, float]
    char: pd.DataFrame
    profiles: pd.DataFrame

    def char_val(self, tech: str, col: str, default: float = 0.0) -> float:
        try:
            v = self.char.at[tech, col]
        except KeyError:
            return default
        if pd.isna(v):
            return default
        try:
            return float(v)
        except (TypeError, ValueError):
            return default

    def must_run_pct(self, tech: str, month: int) -> float:
        """Must-run share for the given 0-based month, as % of installed capacity.

        The column holds either a scalar or a comma-separated 12-value string.
        ("Must Run (Number of units)" is a separate, unreliable column in this
        data -- it is 0 for fleets that "Must Run (%)" shows as partially
        must-run, so the floor is computed from the %-of-capacity column.)
        """
        try:
            raw = self.char.at[tech, "Must Run (%)"]
        except KeyError:
            return 0.0
        return _month_value(raw, month)


def _month_value(raw, month: int) -> float:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    parts = [p.strip() for p in str(raw).split(",") if p.strip() != ""]
    if not parts:
        return 0.0
    idx = min(month, len(parts) - 1)
    try:
        return float(parts[idx])
    except ValueError:
        return 0.0


def _read_kv_ws(ws) -> dict[str, float]:
    """Read a two-column key/value worksheet (skips the 'Code' row)."""
    out: dict[str, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        k, v = row[0], (row[1] if len(row) > 1 else None)
        if not isinstance(k, str) or k == "Code":
            continue
        try:
            out[k] = float(v)
        except (TypeError, ValueError):
            out[k] = 0.0
    return out


def _read_table_ws(ws, row_start: int | None = None, row_end: int | None = None) -> pd.DataFrame:
    """Read a worksheet as a DataFrame. Header is row 1; data rows are the given
    1-based half-open [row_start, row_end) window (all rows if unspecified)."""
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    ncol = len(header)
    data = []
    lo = 0 if row_start is None else row_start
    hi = None if row_end is None else row_end
    for i, r in enumerate(rows):
        if i < lo:
            continue
        if hi is not None and i >= hi:
            break
        data.append(list(r)[:ncol] + [None] * (ncol - len(r)))
    return pd.DataFrame(data, columns=header)


def load_zone(code: str, data_dir: Path, hour_start: int, hour_end: int) -> ZoneData:
    path = Path(data_dir) / f"{code}.xlsx"
    if not path.exists():
        raise FileNotFoundError(f"Zone workbook not found: {path}")
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    capacities = _read_kv_ws(wb[S_CAP])
    storage_energy = _read_kv_ws(wb[S_STO])
    reserves = _read_kv_ws(wb[S_RES])
    h2_assets = {k: v for k, v in _read_kv_ws(wb[S_GH]).items() if "(Gas)" not in k}

    char = _read_table_ws(wb[S_CHAR])
    char = char.set_index(char.columns[0])

    profiles = _read_table_ws(wb[S_PROF], hour_start, hour_end).reset_index(drop=True)
    profiles = profiles.drop(columns=[c for c in profiles.columns if c == "Gas Demand Profile"],
                             errors="ignore")

    wb.close()
    return ZoneData(code, capacities, storage_energy, reserves, h2_assets, char, profiles)


def zones_in_db(db_path: Path) -> list[str]:
    """Sorted list of zone codes present in the zones parquet database."""
    df = pd.read_parquet(db_path, columns=["zone"])
    return sorted(df["zone"].unique().tolist())


def _zone_from_db(zdf: pd.DataFrame, code: str, h0: int, h1: int) -> ZoneData:
    """Reconstruct a ZoneData from this zone's slice of the long parquet table."""
    def scalar(section):
        d = zdf[zdf["section"] == section]
        return dict(zip(d["item"], d["value_num"]))

    c = zdf[zdf["section"] == "characteristics"].copy()
    ti = c["item"].str.split("||", n=1, expand=True, regex=False)
    c["tech"], c["attr"] = ti[0], ti[1]
    c["value"] = c["value_str"].where(c["value_str"].notna(), c["value_num"])
    char = c.pivot(index="tech", columns="attr", values="value")
    char.index.name = "Technology"
    char.columns.name = None

    p = zdf[zdf["section"] == "profiles"]
    prof = p.pivot(index="hour", columns="item", values="value_num")
    prof.columns.name = None
    prof = prof.iloc[h0:h1].reset_index(drop=True)

    return ZoneData(code, scalar("capacities"), scalar("storage_energy"),
                    scalar("reserves"), scalar("h2_assets"), char, prof)


def load_zones_from_db(codes: list[str], db_path: Path,
                       hour_start: int, hour_end: int) -> dict[str, ZoneData]:
    """Load ZoneData for the given zones from ``zones_2030.parquet`` (one read)."""
    db_path = Path(db_path)
    if not db_path.exists():
        raise FileNotFoundError(
            f"Zone database not found: {db_path}. Build it with `python build_zones_db.py`.")
    db = pd.read_parquet(db_path, filters=[("zone", "in", list(codes))])
    present = set(db["zone"].unique())
    missing = [z for z in codes if z not in present]
    if missing:
        raise KeyError(f"zones not in {db_path.name}: {missing}")
    by_zone = {z: g for z, g in db.groupby("zone", sort=False)}
    return {z: _zone_from_db(by_zone[z], z, hour_start, hour_end) for z in codes}


def classify(tech: str) -> tuple[str, bool]:
    """Map a Technology-Capacities row name to (category, is_h2_fuel).

    ``is_h2_fuel`` marks committable plants that consume hydrogen drawn from
    the H2 balance rather than an exogenous fuel. Hydrogen (fc)/(ccgt) are
    NOT flagged this way: their "Fuel (EUR/MWh)" is a real, priced fuel input
    (like natural gas for a Gas plant) -- not sourced from the modelled H2
    network -- so they are priced and dispatched exactly like any other
    thermal fleet (fuel/CO2 cost divided by efficiency, must-run floor, ramp).
    """
    t = tech
    if (t.startswith("Nuclear") or t.startswith("Hard Coal") or t.startswith("Lignite")
            or t.startswith("Gas (") or t.startswith("Light Oil")
            or t.startswith("Heavy oil") or t.startswith("Oil shale")
            or t.startswith("Hydrogen (fc)") or t.startswith("Hydrogen (ccgt)")):
        return CAT_COMMIT, False
    if t.startswith("Wind (") or t.startswith("Solar ("):
        return CAT_VRES, False
    if t.startswith("Hydro (river)"):
        return CAT_ROR, False
    if t.startswith("Other RES") or t.startswith("Other Non-RES") or t.startswith("DSR"):
        return CAT_PROFILE, False
    return CAT_IGNORE, False


VRES_PROFILE = {
    "Wind (onshore) (MW)": "Wind_Onshore Profile",
    "Wind (offshore) (MW)": "Wind_Offshore Profile",
    "Solar (MW)": "Solar Profile",
    "Solar (rooftop) (MW)": "Solar_Rooftop Profile",
    "Solar (thermal) (MW)": "CSP_noStorage Profile",
    "Solar (thermal_with_storage) (MW)": "CSP_withStorage_D Profile",
}


def profile_gen_column(tech: str) -> str:
    """Profile column (MW series) for an Other RES / Other Non-RES / DSR tech."""
    return tech.replace("(MW)", "(MW/h)")
