"""Per-zone hourly series from the PLEXOS reference output that this model
can't derive on its own: marginal prices (for pricing cross-border trade
against each neighbour's own realized marginal cost -- model.py's
priced/controllable import & export legs, without needing a full joint
multi-zone solve) and the "Demand Side Response Implicit" / "Electrolyser
(load)" adjustments used to build each zone's net demand target the same way
the single-zone DE00 validation runs did.

Everything is extracted from the PLEXOS MMStandardOutputFile workbook
(``XLSXs/MMStandardOutputFile_*.xlsx``) in one pass per sheet (electricity:
``Hourly Market Data``; hydrogen: ``Hourly H2 Data``), and written as wide
parquet databases (hour-indexed rows, one column per zone/country) to
``inputs/`` -- the same convention already used by
``inputs/crossborder_electricity_2030.parquet`` /
``crossborder_hydrogen_2030.parquet`` (loaded by exports_loader.py), so a
neighbour's price/adjustment and its transfer volume both come from
consistently-shaped databases.

The workbook itself is only needed to (re)build these files -- once built,
the rest of the pipeline runs entirely from ``inputs/``, same as
``build_db.py``'s zones/networks databases.
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd

from .config import (
    DEFAULT_PLEXOS_REF, DEFAULT_EXPORTS_DIR, DEFAULT_MARGINAL_PRICE_ELEC_DB,
    DEFAULT_MARGINAL_PRICE_H2_DB, HOURS_PER_YEAR,
)

DEFAULT_DSR_IMPLICIT_DB = DEFAULT_EXPORTS_DIR / "dsr_implicit_electricity_2030.parquet"
DEFAULT_ELECTROLYSER_LOAD_DB = DEFAULT_EXPORTS_DIR / "electrolyser_load_electricity_2030.parquet"
DEFAULT_WIND_ONSHORE_DB = DEFAULT_EXPORTS_DIR / "wind_onshore_electricity_2030.parquet"
DEFAULT_WIND_OFFSHORE_DB = DEFAULT_EXPORTS_DIR / "wind_offshore_electricity_2030.parquet"
DEFAULT_ROR_DB = DEFAULT_EXPORTS_DIR / "ror_electricity_2030.parquet"
DEFAULT_SOLAR_PV_DB = DEFAULT_EXPORTS_DIR / "solar_pv_electricity_2030.parquet"
DEFAULT_SOLAR_THERMAL_DB = DEFAULT_EXPORTS_DIR / "solar_thermal_electricity_2030.parquet"
DEFAULT_OTHER_RES_DB = DEFAULT_EXPORTS_DIR / "other_res_electricity_2030.parquet"
DEFAULT_HYDRO_RESERVOIR_DB = DEFAULT_EXPORTS_DIR / "hydro_reservoir_electricity_2030.parquet"
DEFAULT_HYDRO_PONDAGE_DB = DEFAULT_EXPORTS_DIR / "hydro_pondage_electricity_2030.parquet"
DEFAULT_HYDRO_OPEN_PS_DB = DEFAULT_EXPORTS_DIR / "hydro_open_ps_electricity_2030.parquet"
DEFAULT_ELECTROLYSER_GEN_H2_DB = DEFAULT_EXPORTS_DIR / "electrolyser_gen_hydrogen_2030.parquet"

_ELEC_SHEET = ("Hourly Market Data", 11, 12, 14)
_H2_SHEET = ("Hourly H2 Data", 11, 12, 14)


def _extract_sheet(ws, cat_row_i: int, country_row_i: int, data_row0: int,
                   labels: dict[str, str], n_hours: int) -> dict[str, pd.DataFrame]:
    """One pass over the sheet's full row width, pulling out every column
    whose Category cell starts with any of ``labels.values()`` at once --
    {label_key: DataFrame(hour x zone/country)}. Scanning is dominated by
    sheet width, not the number of labels requested, so grabbing several
    series here is nearly free compared to a second full pass."""
    cat_row = next(ws.iter_rows(min_row=cat_row_i, max_row=cat_row_i, values_only=True))
    country_row = next(ws.iter_rows(min_row=country_row_i, max_row=country_row_i, values_only=True))

    col_map: dict[int, tuple[str, str]] = {}
    for i, (cat, country) in enumerate(zip(cat_row, country_row)):
        if not isinstance(cat, str):
            continue
        for key, label in labels.items():
            if cat.startswith(label):
                col_map[i + 1] = (key, country)
                break
    if not col_map:
        return {key: pd.DataFrame() for key in labels}
    max_c = max(col_map)

    out: dict[str, dict[str, np.ndarray]] = {key: {} for key in labels}
    for key, name in col_map.values():
        out[key][name] = np.zeros(n_hours)
    for i, row in enumerate(ws.iter_rows(min_row=data_row0, max_row=data_row0 + n_hours - 1,
                                         min_col=1, max_col=max_c, values_only=True)):
        for col, (key, name) in col_map.items():
            v = row[col - 1]
            out[key][name][i] = 0.0 if v is None else float(v)

    result = {}
    for key, series in out.items():
        df = pd.DataFrame(series)
        df.index.name = "hour"
        result[key] = df
    return result


def build_marginal_price_db(ref_path: Path = DEFAULT_PLEXOS_REF,
                            out_elec: Path = DEFAULT_MARGINAL_PRICE_ELEC_DB,
                            out_h2: Path = DEFAULT_MARGINAL_PRICE_H2_DB,
                            out_dsr_implicit: Path = DEFAULT_DSR_IMPLICIT_DB,
                            out_electrolyser: Path = DEFAULT_ELECTROLYSER_LOAD_DB,
                            out_wind_onshore: Path = DEFAULT_WIND_ONSHORE_DB,
                            out_wind_offshore: Path = DEFAULT_WIND_OFFSHORE_DB,
                            out_ror: Path = DEFAULT_ROR_DB,
                            out_solar_pv: Path = DEFAULT_SOLAR_PV_DB,
                            out_solar_thermal: Path = DEFAULT_SOLAR_THERMAL_DB,
                            out_other_res: Path = DEFAULT_OTHER_RES_DB,
                            out_hydro_reservoir: Path = DEFAULT_HYDRO_RESERVOIR_DB,
                            out_hydro_pondage: Path = DEFAULT_HYDRO_PONDAGE_DB,
                            out_hydro_open_ps: Path = DEFAULT_HYDRO_OPEN_PS_DB,
                            out_electrolyser_gen_h2: Path = DEFAULT_ELECTROLYSER_GEN_H2_DB,
                            n_hours: int = HOURS_PER_YEAR) -> dict[str, Path]:
    """Extract every zone's/country's hourly marginal price (electricity +
    hydrogen), Demand Side Response Implicit, Electrolyser (load), every
    renewable generation category (wind on/offshore, run-of-river, solar PV,
    solar thermal, other renewables), and every natural-inflow hydro storage
    category (reservoir, pondage, open-loop pumped) from the PLEXOS reference
    workbook, and write them as wide parquet databases. Slow-ish (a couple
    minutes: the electricity sheet is ~4,200 columns wide, so the single pass
    over it dominates) -- a one-time build step, not part of the model's
    runtime hot path."""
    import openpyxl
    wb = openpyxl.load_workbook(ref_path, read_only=True, data_only=True)
    written = {}
    try:
        sheet, cat_i, country_i, data0 = _ELEC_SHEET
        print(f"Extracting electricity series from '{sheet}'...")
        elec = _extract_sheet(wb[sheet], cat_i, country_i, data0, {
            "price": "Marginal Cost",
            "dsr_implicit": "Demand Side Response Implicit",
            "electrolyser": "Electrolyser (load)",
            "wind_onshore": "Wind Onshore",
            "wind_offshore": "Wind Offshore",
            "ror": "Run-of-River",
            "solar_pv": "Solar (Photovoltaic)",
            "solar_thermal": "Solar (Thermal)",
            "other_res": "Others renewable",
            "hydro_reservoir": "Reservoir",
            "hydro_pondage": "Pondage",
            "hydro_open_ps": "Pump Storage - Open Loop (turbine)",
        }, n_hours)
        for key, out_path in [("price", out_elec), ("dsr_implicit", out_dsr_implicit),
                              ("electrolyser", out_electrolyser),
                              ("wind_onshore", out_wind_onshore), ("wind_offshore", out_wind_offshore),
                              ("ror", out_ror), ("solar_pv", out_solar_pv),
                              ("solar_thermal", out_solar_thermal), ("other_res", out_other_res),
                              ("hydro_reservoir", out_hydro_reservoir),
                              ("hydro_pondage", out_hydro_pondage),
                              ("hydro_open_ps", out_hydro_open_ps)]:
            df = elec[key]
            out_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_parquet(out_path, compression="zstd")
            print(f"  wrote {out_path.name}  ({df.shape[0]} hours x {df.shape[1]} zones)")
            written[key] = out_path

        sheet, cat_i, country_i, data0 = _H2_SHEET
        print(f"Extracting hydrogen series from '{sheet}'...")
        h2 = _extract_sheet(wb[sheet], cat_i, country_i, data0, {
            "price": "Marginal Cost",
            "electrolyser_gen": "Electrolyser (gen.)",
        }, n_hours)
        for key, out_path in [("price", out_h2), ("electrolyser_gen", out_electrolyser_gen_h2)]:
            df = h2[key]
            out_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_parquet(out_path, compression="zstd")
            print(f"  wrote {out_path.name}  ({df.shape[0]} hours x {df.shape[1]} countries)")
            written[key if key != "price" else "price_h2"] = out_path
    finally:
        wb.close()
    return written


def load_zone_series(names: list[str], hours: pd.Index,
                     db_path: Path = DEFAULT_MARGINAL_PRICE_ELEC_DB) -> pd.DataFrame:
    """Read back a subset of zones/countries and an hour window from any of
    the wide parquet databases built by ``build_marginal_price_db`` (pass
    ``db_path=DEFAULT_MARGINAL_PRICE_H2_DB`` for hydrogen prices,
    ``DEFAULT_DSR_IMPLICIT_DB`` / ``DEFAULT_ELECTROLYSER_LOAD_DB`` for the
    demand adjustments). Missing names come back as all-zero columns (e.g. a
    zone genuinely outside PLEXOS's own simulated area) rather than raising."""
    df = pd.read_parquet(db_path)
    h0, h1 = int(hours[0]), int(hours[-1]) + 1
    out = pd.DataFrame(index=hours)
    for n in names:
        out[n] = df[n].to_numpy()[h0:h1] if n in df.columns else np.zeros(len(hours))
    return out


load_marginal_prices = load_zone_series
